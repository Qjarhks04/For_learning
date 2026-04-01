from openpyxl import load_workbook, Workbook 
#openpyxl에서 load_workbook, Workbook을 불러온다
#파이썬에서 엑셀을 사용하기 위한 라이브러리 입니다

user_name = 1 #변수에 1을 넣어 선언합니다 유저이름
user_id = 2 #변수에 2를 넣어 선언 유저아이디
user_money = 3 #변수에 3을 넣어 선언 유저머니
user_lv = 4 #변수에 4를 넣어 선언 유저레벨

default_money = 50000 # 초기 보유머니 설정

#wb변수에 경로에 있는 userDB.xlsx라는 엑셀파일을 불러온다
wb = load_workbook("/Users/jeongbeomgwan/Desktop/coding/pythonpj/userDB.xlsx")
#ws변수에 불러온 wb변수를 활성화하고 넣어준다
ws = wb.active

def loadFile(): #함수 선언 파일을 불러오는 함수
    wb = load_workbook("/Users/jeongbeomgwan/Desktop/coding/pythonpj/userDB.xlsx")
    ws = wb.active
def saveFile(): #함수 선언 엑셀파일을 세이브하고 닫는 함수
    wb.save("/Users/jeongbeomgwan/Desktop/coding/pythonpj/userDB.xlsx")
    wb.close()

#DB(엑셀) 안에 값이 무슨 행까지 있는지 값을 알아내는 함수
def checkRow():
    loadFile() #함수호출
    #2부터 워크시트의 마지막행 번호의 +1이므로 마지막행 번호까지 반복하고 
    #row에 2부터 ws.max_row의 값이 반복하며 들어감
    for row in range(2, ws.max_row + 1):
        #만약 워크시크의 row행, 1열의 값이 None 즉 값이 없을 경우 row값을 return하고 break한다
        if ws.cell(row,1).value is None:
            return row
            break
    #_result에 ws.max_row+1의 값이 들어간다
    _result = ws.max_row+1 
    saveFile() #함수호출
    return _result #_result를 return한다

#DB(엑셀)에 들어가 있는 유저의 개수를 알아내는 함수
def checkUserNum(): #함수 선언
    print("user.py - checkUserNum")
    loadFile() #함수호출

    count = 0 #count라는 변수를 선언하고 0으로 초기화합니다

    #2부터 워크시트의 마지막행 번호의 +1이므로 마지막행 번호까지 반복하고 
    #row에 2부터 ws.max_row의 값이 반복하며 들어감
    for row in range(2, ws.max_row+1):
        #row행, user_name(1)열의 값이 None 즉 값이 있을 경우
        #count에 1씩 추가해준다
        if ws.cell(row,user_name).value != None:
            count = count+1
            #count += 1
        else: #값이 없는경우 count값을 그대로 둔다.
            count = count
    return count #count를 return한다.

#DB의 특정 유저가 있는지 확인하는 함수
def checkUser(_name, _id):
    print("user.py - checkUser")
    #받아온 유저의 이름과 아이디를 스트링 형태로 출력한다 터미널에
    print(str(_name) + "<" + str(_id) + ">의 존재 여부 확인")
    print("")

    loadFile() #함수 호출

    userNum = checkUserNum() #checkUserNum()의 함수에서 리턴된 값을 userNum이라는 변수에 넣어둔다
    print("등록된 유저수: ", userNum) #userNum의 값을 터미널에 출력한다
    print("")

    print("이름과 고유번호 탐색")
    print("")

    for row in range(2, 3+userNum): #2부터 3+userNum의 값만큼 반복하고 row에 값을 넣어준다
        #row값과 워크시트의 row행, user_name(1)열의 값을 출력
        print(row, "번째 줄 name: ", ws.cell(row,user_name).value)
        #Discordbot.py에서 받아온 유저이름을 출력
        print("입력된 name: ", _name)
        #위에 2개의 값을 비교하여 일치 여부를 확인하고 출력
        print("이름과 일치 여부: ", ws.cell(row, user_name).value == _name)

        #위랑 비슷하게 값을 받아온다 user_id(2)이다
        print(row,"번째 줄 id: ", ws.cell(row,user_id).value)
        #Discordbot.py에서 받아온 유저아이디를 출력
        print("입력된 id: ", _id)
        #위에 받아온 둘의 값이 같은지 비교하고 출력
        print("고유번호정보와 일치 여부: ", ws.cell(row, user_id).value == _id)
        print("")

        #위에서 비교했던 db의 이름과 유저의 이름, db의 아이디값과 유저의 아이디값을 비교후
        #모두 일치하면 값을 출력하고 True(1), row의 값을 돌려주고 break합니다
        if ws.cell(row, user_name).value == _name and ws.cell(row,user_id).value == _id:
            print("등록된  이름과 고유번호를 발견")
            print("등록된  값의 위치: ",  row, "번째 줄")
            print("")

            saveFile()

            return True, row
            break
        else: #그렇지 않을경우 반복한다
            print("등록된 정보를 탐색 실패, 재탐색 실시")

    saveFile()
    print("발견 실패")
    #마지막 반복까지 값을 찾지 못하는 경우 False(0), None(값이 없음)을 리턴한다
    return False, None

#특정 유저가 보유한 돈이 얼마나 있는지 확인하는 함수
def getMoney(_name,_row):
    print("user.py - getMoney")
    loadFile()

    print(_name, "의 돈을 탐색")
    #result의 변수에 워크시트의 _row행, user_money(3)열의 값을 가져와 넣어준다
    result = ws.cell(_row, user_money).value
    print(_name,"의 보유 자산: ", result)

    saveFile()
    #result를 return합니다
    return result

#특정 유저의 보유 금액을 수정하는 함수
def modifyMoney(_target, _row, _amount):
    print("user.py - modifyMoney")
    loadFile()

    print(_target, "의 자산데이터 수정")
    print(_target, "의 자산: " + str(ws.cell(_row, user_money).value))
    print("추가할 액수: ", _amount)
    #_row행, user_money(3)열의 값에 _amount값을 추가한다
    ws.cell(_row, user_money).value += _amount

    print("자산데이터 수정 완료")
    print("수정된", _target, "의 자산: ", ws.cell(_row, user_money).value)
    
    saveFile()

#송금할때 사용하는 함수
def remit(sender, sender_row, receiver, receiver_row, _amount):
    print("user.py - remit")
    print("보내는 사람: ", sender)
    print("받는 사람: ", receiver)
    print("보내는 돈: ", _amount)
    print("")

    #돈을 받아 돈을 추가하기 위하여 사용하는 함수
    #receiver은 받는 사람
    modifyMoney(receiver, receiver_row, int(_amount))
    #돈을 보내서 돈을 줄이기 위하여 사용하는 함수
    #sender는 보내는 사람
    modifyMoney(sender, sender_row, -int(_amount))

    print("")

#회원가입 함수
def signup(_name, _id):
    loadFile()
    #_row에 checkRow()함수에서 나온 값을 넣어준다
    #_row에는 DB를 체크하여 값이 없는 행의 값입니다
    _row = checkRow()
    #밑에 4줄 코드는 _row행, column은 밑의 변수의 값에 해당하는 열에 value에 해당하는 값을 넣어주는 코드입니다
    ws.cell(row=_row, column=user_name, value=_name)
    ws.cell(row=_row, column=user_id, value=_id)
    ws.cell(row=_row, column=user_money, value=default_money)
    ws.cell(row=_row, column=user_lv, value=1)
    saveFile()

def delete():
    loadFile() #함수호출
    print("유저 데이터를 삭제") #터미널에 출력
    ws.delete_rows(2, ws.max_row) #워크시트의 2행부터 마지막 행까지의 모든 정보를 지운다
    saveFile() #함수호출

def userInfo(_row):
    loadFile() #함수호출

    #_lv에 _row행, user_lv열의 값을 넣는다
    _lv = ws.cell(_row,user_lv).value
    #_money에 _row행, user_money열의 값을 넣는다
    _money = ws.cell(_row,user_money).value

    print("레벨: ", _lv)
    print("보유자산: ", _money)

    saveFile()

    #_lv, _money의 값을 리턴한다
    return _lv, _money